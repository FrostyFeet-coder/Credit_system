import datetime
import openpyxl
from celery import shared_task
from .models import Customer, Loan
from .models import Loan

@shared_task
def process_customer_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            # row[1] = First Name
            # row[2] = Last Name
            # row[3] = Age
            # row[4] = Phone Number
            # row[5] = Monthly Salary
            # row[6] = Approved Limit

            Customer.objects.create(
                first_name=row[1],
                last_name=row[2],
                age=int(row[3]),
                phone_number=str(row[4]),
                monthly_income=float(row[5]),
                approved_limit=float(row[6]),
                current_debt=0  # default as per business logic
            )
        except Exception as e:
            print(f"[ERROR] Skipping row: {row} → {e}")

@shared_task
def process_loan_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            if not any(row):
                continue

            customer_id = int(row[0])
            loan_amount = float(row[2])
            tenure = int(row[3])
            interest_rate = float(row[4])
            monthly_installment = float(row[5])
            emis_paid_on_time = int(row[6])

            # Dates are already datetime.datetime → just extract .date()
            approval_date = row[7].date()
            end_date = row[8].date()

            Loan.objects.create(
                customer_id=customer_id,
                loan_amount=loan_amount,
                tenure=tenure,
                interest_rate=interest_rate,
                monthly_installment=monthly_installment,
                emis_paid_on_time=emis_paid_on_time,
                start_date=approval_date,
                end_date=end_date
            )
        except Exception as e:
            print(f"[ERROR] Skipping row: {row} → {e}")
